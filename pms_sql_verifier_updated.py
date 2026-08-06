#!/usr/bin/env python3
"""
PMS SQL Query Verifier - Updated for the Reorganized Executive/KPI Workbook
===========================================================================

Purpose
-------
Batch-validate only the SQL queries that are already mapped in:
    PMS_Questionnaire_Reorganized_Executive_KPI.xlsx

The verifier does NOT attempt to generate SQL for pending executive/KPI questions.
It validates syntax, table/column references, parameter binding, execution, returned
columns, and whether representative data is returned.

Safety
------
- Uses a read-only database user.
- Starts a READ ONLY transaction when supported.
- Accepts only SELECT/WITH statements.
- Does not append LIMIT to SQL (which can break UNION/CTE queries).
- Fetches only a small sample from the result cursor.
- Blocks queries that reference tables listed as empty in the workbook.

Usage
-----
1. Install dependencies:
       python -m pip install pymysql openpyxl

2. Place these files in the same folder:
       pms_sql_verifier_updated.py
       PMS_Questionnaire_Reorganized_Executive_KPI.xlsx
       .env

3. Example .env:
       MYSQL_HOST=127.0.0.1
       MYSQL_PORT=3306
       MYSQL_USER=pms_readonly
       MYSQL_PASSWORD=change_me
       MYSQL_DATABASE=engagedb

       # Optional overrides
       PMS_CURRENT_USER_ID=3066
       PMS_PROJECT_NAME=Real Project Name
       PMS_PROJECT_IDENTIFIER=PF-001
       PMS_TASK_REFERENCE=TASK-001
       PMS_RISK_REFERENCE=RISK-001
       PMS_ISSUE_REFERENCE=ISS-001
       PMS_EMPLOYEE_NAME=Rahul Sharma
       PMS_DEPARTMENT_NAME=Engineering
       PMS_PROJECT_TYPE=Fixed Price
       PMS_START_DATE=2026-01-01
       PMS_END_DATE=2026-08-04
       PMS_DAYS=30
       PMS_RUN_EXPLAIN=1
       PMS_CAPTURE_SAMPLE_DATA=0

4. Run:
       python pms_sql_verifier_updated.py

5. Output:
       pms_verification_results.json

MCP usage
---------
This script connects directly with PyMySQL. It does not call an MCP server.
Use your local DB MCP separately for:
- confirming business-rule mappings and lookup meanings;
- inspecting representative records;
- running/iterating complex KPI SQL;
- checking EXPLAIN plans;
- validating whether results make business sense.
"""

from __future__ import annotations

import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "PMS_Questionnaire_Reorganized_Executive_KPI.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "pms_verification_results.json"
DEFAULT_SHEET = "Reorganized Questionnaire"
EMPTY_TABLE_SHEET = "Empty Table"

MAPPED_STATUS_PREFIXES = (
    "Mapped -",
    "Mapped –",  # en dash used in workbook
)


def load_dotenv(path: Path) -> None:
    """Load simple KEY=VALUE pairs without overwriting existing environment variables."""
    if not path.exists():
        return
    with path.open("r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("export "):
                line = line[7:].strip()
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key:
                os.environ.setdefault(key, value)


load_dotenv(SCRIPT_DIR / ".env")


def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def env_int(name: str, default: int | None = None) -> int | None:
    value = os.getenv(name)
    if value is None or value.strip() == "":
        return default
    return int(value)


@dataclass
class QueryDefinition:
    source_row: int
    original_id: Any
    canonical: str
    question: str
    mapping_status: str
    data_readiness: str
    sql: str
    bind_params: list[str]
    referenced_tables: list[str]


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def is_mapped_status(status: str) -> bool:
    return any(status.startswith(prefix) for prefix in MAPPED_STATUS_PREFIXES)


def extract_bind_params(sql: str) -> list[str]:
    found = re.findall(r":([A-Za-z_][A-Za-z0-9_]*)\b", sql)
    return list(dict.fromkeys(found))


def extract_referenced_tables(sql: str) -> list[str]:
    tables = re.findall(
        r"\b(?:FROM|JOIN)\s+`?([A-Za-z_][A-Za-z0-9_]*)`?",
        sql,
        flags=re.IGNORECASE,
    )
    return sorted(set(tables), key=str.lower)


def load_empty_tables(workbook_path: Path) -> set[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if EMPTY_TABLE_SHEET not in workbook.sheetnames:
        return set()

    sheet = workbook[EMPTY_TABLE_SHEET]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return set()

    header_map = {normalize_header(value): index for index, value in enumerate(header)}
    table_index = header_map.get("Table Name")
    if table_index is None:
        return set()

    result: set[str] = set()
    for row in rows:
        if table_index < len(row) and row[table_index]:
            result.add(str(row[table_index]).strip().lower())
    return result


def load_mapped_queries(workbook_path: Path, sheet_name: str = DEFAULT_SHEET) -> list[QueryDefinition]:
    """Load mapped SQL rows using workbook header names rather than fixed column positions."""
    from openpyxl import load_workbook

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ValueError(f"Sheet '{sheet_name}' is empty") from exc

    headers = {normalize_header(value): index for index, value in enumerate(header) if value is not None}

    required_headers = {
        "Canonical Intent",
        "Primary Question",
        "Query Mapping Status",
        "SQL Query (Existing or Recovered)",
    }
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        raise ValueError(f"Missing expected columns in '{sheet_name}': {', '.join(missing_headers)}")

    def value(row: tuple[Any, ...], header_name: str) -> Any:
        index = headers.get(header_name)
        return row[index] if index is not None and index < len(row) else None

    queries: list[QueryDefinition] = []
    for excel_row, row in enumerate(rows, start=2):
        sql = str(value(row, "SQL Query (Existing or Recovered)") or "").strip()
        mapping_status = str(value(row, "Query Mapping Status") or "").strip()
        data_readiness = str(value(row, "Data Readiness") or "").strip()

        # Only test rows deliberately marked as mapped and containing SQL.
        if not sql or not is_mapped_status(mapping_status):
            continue

        queries.append(
            QueryDefinition(
                source_row=excel_row,
                original_id=value(row, "Original ID"),
                canonical=str(value(row, "Canonical Intent") or "").strip(),
                question=str(value(row, "Primary Question") or "").strip(),
                mapping_status=mapping_status,
                data_readiness=data_readiness,
                sql=sql.rstrip().rstrip(";"),
                bind_params=extract_bind_params(sql),
                referenced_tables=extract_referenced_tables(sql),
            )
        )

    return queries


DB_CONFIG = {
    "host": os.getenv("MYSQL_HOST", ""),
    "port": int(os.getenv("MYSQL_PORT", "3306")),
    "user": os.getenv("MYSQL_USER", ""),
    "password": os.getenv("MYSQL_PASSWORD", ""),
    "database": os.getenv("MYSQL_DATABASE", ""),
    "charset": "utf8mb4",
    "connect_timeout": int(os.getenv("MYSQL_CONNECT_TIMEOUT", "10")),
    "read_timeout": int(os.getenv("MYSQL_READ_TIMEOUT", "60")),
    "write_timeout": int(os.getenv("MYSQL_WRITE_TIMEOUT", "10")),
    "autocommit": False,
}


TEST_PARAMS: dict[str, Any] = {
    "current_user_id": env_int("PMS_CURRENT_USER_ID"),
    "project_name": os.getenv("PMS_PROJECT_NAME"),
    "project_identifier": os.getenv("PMS_PROJECT_IDENTIFIER"),
    "task_reference": os.getenv("PMS_TASK_REFERENCE"),
    "risk_reference": os.getenv("PMS_RISK_REFERENCE"),
    "issue_reference": os.getenv("PMS_ISSUE_REFERENCE"),
    "employee_name": os.getenv("PMS_EMPLOYEE_NAME"),
    "department_name": os.getenv("PMS_DEPARTMENT_NAME"),
    "project_type": os.getenv("PMS_PROJECT_TYPE"),
    "start_date": os.getenv("PMS_START_DATE"),
    "end_date": os.getenv("PMS_END_DATE"),
    "days": env_int("PMS_DAYS", 30),
}

MAX_SAMPLE_ROWS = int(os.getenv("PMS_SAMPLE_ROWS", "5"))
CAPTURE_SAMPLE_DATA = env_bool("PMS_CAPTURE_SAMPLE_DATA", False)
RUN_EXPLAIN = env_bool("PMS_RUN_EXPLAIN", True)
AUTO_DISCOVER_PARAMS = env_bool("PMS_AUTO_DISCOVER_PARAMS", True)


def first_scalar(cursor, sql: str) -> Any:
    cursor.execute(sql)
    row = cursor.fetchone()
    return row[0] if row else None


def discover_test_params(cursor) -> dict[str, Any]:
    """Find representative values only for missing test parameters."""
    discovered: dict[str, Any] = {}

    discovery_queries: dict[str, str] = {
        "current_user_id": """
            SELECT pu.user_id
            FROM tbl_project_users pu
            JOIN tbl_users u ON u.user_id = pu.user_id
            WHERE pu.status = 1 AND pu.request_status = 'A'
            ORDER BY pu.user_id
            LIMIT 1
        """,
        "project_name": """
            SELECT p.title
            FROM tbl_projects p
            WHERE p.status = 1 AND p.archive = 0 AND p.title IS NOT NULL AND p.title <> ''
            ORDER BY p.project_id DESC
            LIMIT 1
        """,
        "project_identifier": """
            SELECT COALESCE(NULLIF(p.project_unique_id, ''), NULLIF(p.project_key, ''), p.title)
            FROM tbl_projects p
            WHERE p.status = 1 AND p.archive = 0
            ORDER BY p.project_id DESC
            LIMIT 1
        """,
        "task_reference": """
            SELECT t.reference_no
            FROM tbl_project_tasks t
            WHERE t.status = 1 AND t.archive = 0
              AND t.reference_no IS NOT NULL AND t.reference_no <> ''
            ORDER BY t.task_id DESC
            LIMIT 1
        """,
        "risk_reference": """
            SELECT r.reference_no
            FROM tbl_risk r
            WHERE r.status = 1 AND r.reference_no IS NOT NULL AND r.reference_no <> ''
            ORDER BY r.risk_id DESC
            LIMIT 1
        """,
        "issue_reference": """
            SELECT i.reference_no
            FROM tbl_project_issues i
            WHERE i.status = 1 AND i.reference_no IS NOT NULL AND i.reference_no <> ''
            ORDER BY i.issue_id DESC
            LIMIT 1
        """,
        "employee_name": """
            SELECT CONCAT_WS(' ', u.first_name, u.last_name)
            FROM tbl_users u
            WHERE u.first_name IS NOT NULL
            ORDER BY u.user_id DESC
            LIMIT 1
        """,
        "department_name": """
            SELECT d.department_name
            FROM department d
            WHERE d.department_name IS NOT NULL AND d.department_name <> ''
            ORDER BY d.department_id
            LIMIT 1
        """,
        "project_type": """
            SELECT pt.project_type
            FROM tbl_project_types pt
            WHERE pt.status = 1 AND pt.project_type IS NOT NULL AND pt.project_type <> ''
            ORDER BY pt.project_type_id
            LIMIT 1
        """,
        "start_date": """
            SELECT DATE_FORMAT(COALESCE(MIN(ts.timesheet_date), DATE_SUB(CURRENT_DATE, INTERVAL 90 DAY)), '%Y-%m-%d')
            FROM tbl_timesheets ts
        """,
        "end_date": "SELECT DATE_FORMAT(CURRENT_DATE, '%Y-%m-%d')",
    }

    for key, sql in discovery_queries.items():
        if TEST_PARAMS.get(key) is not None:
            continue
        try:
            discovered[key] = first_scalar(cursor, sql)
        except Exception as exc:  # Discovery failure should not abort query verification.
            discovered[key] = None
            print(f"  Auto-discovery warning for {key}: {exc}")

    return discovered


def replace_bind_params(sql: str, params: Iterable[str]) -> str:
    result = sql
    for param in sorted(params, key=len, reverse=True):
        result = re.sub(rf":{re.escape(param)}\b", f"%({param})s", result)
    return result


def remove_leading_comments(sql: str) -> str:
    text = re.sub(r"^\s*(?:--[^\n]*\n\s*)+", "", sql)
    text = re.sub(r"^\s*/\*.*?\*/\s*", "", text, flags=re.DOTALL)
    return text.strip()


def validate_read_only_sql(sql: str) -> None:
    clean = remove_leading_comments(sql).rstrip().rstrip(";").strip()
    if not re.match(r"^(SELECT|WITH)\b", clean, flags=re.IGNORECASE):
        raise ValueError("Only SELECT or WITH queries are allowed")

    # Reject multiple statements. A single trailing semicolon was removed above.
    if ";" in clean:
        raise ValueError("Multiple SQL statements are not allowed")

    forbidden = re.compile(
        r"\b(INSERT|UPDATE|DELETE|REPLACE|ALTER|DROP|TRUNCATE|CREATE|GRANT|REVOKE|CALL|LOAD|INTO\s+OUTFILE|INTO\s+DUMPFILE)\b",
        flags=re.IGNORECASE,
    )
    if forbidden.search(clean):
        raise ValueError("Potential write or administrative statement detected")


def serialize_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, (int, float, str, bool)) or value is None:
        return value
    return str(value)


def sanitize_sample(columns: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    """Mask common PII columns in optional sample output."""
    record: dict[str, Any] = {}
    for column, raw_value in zip(columns, row):
        value = serialize_value(raw_value)
        lower = column.lower()
        if value is not None and any(token in lower for token in ("email", "phone", "mobile")):
            record[column] = "[MASKED]"
        else:
            record[column] = value
    return record


def execute_explain(cursor, sql: str, params: dict[str, Any]) -> dict[str, Any]:
    try:
        cursor.execute("EXPLAIN " + sql, params)
        columns = [item[0] for item in cursor.description or []]
        rows = cursor.fetchmany(20)
        return {
            "status": "PASS",
            "columns": columns,
            "rows": [sanitize_sample(columns, row) for row in rows],
        }
    except Exception as exc:
        return {"status": "FAIL", "error": f"{type(exc).__name__}: {exc}"}


def verify_query(cursor, query: QueryDefinition, test_params: dict[str, Any], empty_tables: set[str]) -> dict[str, Any]:
    result: dict[str, Any] = {
        "workbook_row": query.source_row,
        "original_id": query.original_id,
        "canonical": query.canonical,
        "question": query.question,
        "mapping_status": query.mapping_status,
        "data_readiness": query.data_readiness,
        "referenced_tables": query.referenced_tables,
        "bind_params": query.bind_params,
        "status": None,
        "columns": [],
        "sample_row_count": 0,
        "sample_data": [],
        "error": None,
        "explain": None,
    }

    blocked_tables = sorted(table for table in query.referenced_tables if table.lower() in empty_tables)
    if blocked_tables:
        result["status"] = "BLOCKED_EMPTY_TABLE"
        result["error"] = "References empty table(s): " + ", ".join(blocked_tables)
        return result

    missing = [param for param in query.bind_params if test_params.get(param) is None]
    if missing:
        result["status"] = "SKIPPED_MISSING_PARAMS"
        result["error"] = "Missing test parameters: " + ", ".join(missing)
        return result

    try:
        validate_read_only_sql(query.sql)
    except Exception as exc:
        result["status"] = "BLOCKED_UNSAFE_SQL"
        result["error"] = str(exc)
        return result

    param_values = {param: test_params[param] for param in query.bind_params}
    sql = replace_bind_params(query.sql, query.bind_params)

    if RUN_EXPLAIN:
        result["explain"] = execute_explain(cursor, sql, param_values)

    try:
        cursor.execute(sql, param_values)
        columns = [item[0] for item in cursor.description or []]
        rows = cursor.fetchmany(MAX_SAMPLE_ROWS)

        result["columns"] = columns
        result["sample_row_count"] = len(rows)
        result["status"] = "PASS_WITH_DATA" if rows else "PASS_EMPTY_RESULT"

        if CAPTURE_SAMPLE_DATA:
            result["sample_data"] = [sanitize_sample(columns, row) for row in rows]

    except Exception as exc:
        result["status"] = "FAIL"
        result["error"] = f"{type(exc).__name__}: {exc}"

    return result


LOOKUP_QUERIES: dict[str, str] = {
    "project_statuses_used": """
        SELECT DISTINCT p.project_status_id, ps.status_name, ps.color_code, ps.status_order, ps.status
        FROM tbl_projects p
        LEFT JOIN tbl_project_status ps ON ps.status_id = p.project_status_id
        ORDER BY p.project_status_id
    """,
    "project_stages_used": """
        SELECT DISTINCT p.project_stage_id, st.stage_name, st.color_code, st.status
        FROM tbl_projects p
        LEFT JOIN tbl_project_stages st ON st.stage_id = p.project_stage_id
        ORDER BY p.project_stage_id
    """,
    "project_priority_values": """
        SELECT p.priority, COUNT(*) AS project_count
        FROM tbl_projects p
        GROUP BY p.priority
        ORDER BY p.priority
    """,
    "project_rag_values": """
        SELECT p.rag, COUNT(*) AS project_count
        FROM tbl_projects p
        GROUP BY p.rag
        ORDER BY p.rag
    """,
    "task_statuses_used": """
        SELECT DISTINCT t.task_status_id, ts.status_name, ts.status_type, ts.status_order, ts.status
        FROM tbl_project_tasks t
        LEFT JOIN tbl_project_task_status ts ON ts.task_status_id = t.task_status_id
        ORDER BY t.task_status_id
    """,
    "task_priority_values": """
        SELECT t.priority, COUNT(*) AS task_count
        FROM tbl_project_tasks t
        GROUP BY t.priority
        ORDER BY t.priority
    """,
    "task_rag_values": """
        SELECT t.rag, COUNT(*) AS task_count
        FROM tbl_project_tasks t
        GROUP BY t.rag
        ORDER BY t.rag
    """,
    "issue_statuses_used": """
        SELECT DISTINCT i.issue_status_id, s.status_name, s.status_type, s.status_order, s.status
        FROM tbl_project_issues i
        LEFT JOIN tbl_project_issue_status s ON s.issue_status_id = i.issue_status_id
        ORDER BY i.issue_status_id
    """,
    "issue_severity_values": """
        SELECT i.severity, COUNT(*) AS issue_count
        FROM tbl_project_issues i
        GROUP BY i.severity
        ORDER BY i.severity
    """,
    "issue_priority_values": """
        SELECT i.priority, COUNT(*) AS issue_count
        FROM tbl_project_issues i
        GROUP BY i.priority
        ORDER BY i.priority
    """,
    "risk_statuses_used": """
        SELECT DISTINCT r.status_id, rs.status_name
        FROM tbl_risk r
        LEFT JOIN tbl_risk_status rs ON rs.status_id = r.status_id
        ORDER BY r.status_id
    """,
    "risk_priorities_used": """
        SELECT DISTINCT r.risk_priority_id, rp.priority_name
        FROM tbl_risk r
        LEFT JOIN tbl_risk_priority rp ON rp.risk_priority_id = r.risk_priority_id
        ORDER BY r.risk_priority_id
    """,
    "risk_probabilities_used": """
        SELECT DISTINCT r.probability_id, rp.probability_name, rp.probability_value
        FROM tbl_risk r
        LEFT JOIN tbl_risk_probability rp ON rp.probability_id = r.probability_id
        ORDER BY r.probability_id
    """,
    "risk_impacts_used": """
        SELECT DISTINCT r.impact_id, ri.impact_name, ri.impact_value
        FROM tbl_risk r
        LEFT JOIN tbl_risk_impact ri ON ri.impact_id = r.impact_id
        ORDER BY r.impact_id
    """,
}


def run_lookup_queries(cursor) -> dict[str, Any]:
    results: dict[str, Any] = {}
    for name, sql in LOOKUP_QUERIES.items():
        try:
            validate_read_only_sql(sql)
            cursor.execute(sql)
            columns = [item[0] for item in cursor.description or []]
            rows = cursor.fetchmany(500)
            results[name] = {
                "status": "PASS",
                "count": len(rows),
                "columns": columns,
                "data": [sanitize_sample(columns, row) for row in rows],
            }
            print(f"  {name}: {len(rows)} rows")
        except Exception as exc:
            results[name] = {"status": "FAIL", "error": f"{type(exc).__name__}: {exc}"}
            print(f"  {name}: FAIL - {exc}")
    return results


def main() -> int:
    try:
        import pymysql
    except ImportError:
        print("ERROR: Dependencies missing. Run: python -m pip install pymysql openpyxl")
        return 1

    workbook_path = Path(os.getenv("PMS_WORKBOOK_PATH", str(DEFAULT_WORKBOOK))).expanduser().resolve()
    output_path = Path(os.getenv("PMS_OUTPUT_PATH", str(DEFAULT_OUTPUT))).expanduser().resolve()

    missing_db = [key for key in ("host", "user", "password", "database") if not DB_CONFIG.get(key)]
    if missing_db:
        print("ERROR: Missing database settings: " + ", ".join(missing_db))
        print("Set MYSQL_HOST, MYSQL_USER, MYSQL_PASSWORD, and MYSQL_DATABASE in .env.")
        return 1

    try:
        queries = load_mapped_queries(workbook_path)
        empty_tables = load_empty_tables(workbook_path)
    except Exception as exc:
        print(f"ERROR loading workbook: {exc}")
        return 1

    print(f"Workbook: {workbook_path.name}")
    print(f"Mapped SQL queries selected: {len(queries)}")
    print(f"Known empty tables loaded: {len(empty_tables)}")

    try:
        connection = pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.SSCursor)
    except Exception as exc:
        print(f"CONNECTION FAILED: {exc}")
        return 1

    cursor = connection.cursor()
    read_only_transaction = False
    try:
        try:
            cursor.execute("SET SESSION TRANSACTION READ ONLY")
            cursor.execute("START TRANSACTION READ ONLY")
            read_only_transaction = True
        except Exception as exc:
            # A properly restricted DB account remains the primary safety control.
            print(f"Warning: could not start explicit read-only transaction: {exc}")

        effective_params = dict(TEST_PARAMS)
        if AUTO_DISCOVER_PARAMS:
            print("\nAuto-discovering representative test parameters...")
            discovered = discover_test_params(cursor)
            for key, value in discovered.items():
                if effective_params.get(key) is None and value is not None:
                    effective_params[key] = value
                    print(f"  {key}: discovered")

        print("\nVerifying mapped SQL queries...")
        print(f"{'ID':<8} {'Status':<24} {'Canonical Intent':<42} {'Rows':<5} Error")
        print("-" * 120)

        results: list[dict[str, Any]] = []
        summary: dict[str, int] = {}

        for query in queries:
            result = verify_query(cursor, query, effective_params, empty_tables)
            results.append(result)
            status = str(result["status"])
            summary[status] = summary.get(status, 0) + 1
            error = str(result.get("error") or "")[:45]
            print(
                f"{str(query.original_id or query.source_row):<8} "
                f"{status:<24} "
                f"{query.canonical[:41]:<42} "
                f"{result['sample_row_count']:<5} "
                f"{error}"
            )

        print("\nExtracting lookup values used by the live data...")
        lookup_results = run_lookup_queries(cursor)

        output = {
            "generated_at": datetime.now().isoformat(),
            "workbook": str(workbook_path),
            "sheet": DEFAULT_SHEET,
            "database": DB_CONFIG["database"],
            "read_only_transaction_started": read_only_transaction,
            "summary": {
                "selected_queries": len(queries),
                "status_counts": summary,
            },
            "test_params": {
                key: ("(set)" if value is not None else "(missing)")
                for key, value in effective_params.items()
            },
            "query_results": results,
            "lookup_master_data": lookup_results,
        }

        with output_path.open("w", encoding="utf-8") as handle:
            json.dump(output, handle, indent=2, ensure_ascii=False)

        print("\nSummary:")
        for status, count in sorted(summary.items()):
            print(f"  {status}: {count}")
        print(f"\nResults written to: {output_path}")
        return 0

    finally:
        try:
            connection.rollback()
        except Exception:
            pass
        cursor.close()
        connection.close()


if __name__ == "__main__":
    sys.exit(main())
